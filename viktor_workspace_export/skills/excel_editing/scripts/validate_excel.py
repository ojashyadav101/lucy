#!/usr/bin/env python3
"""
Excel Validation Script

Validates Excel files for formula errors and optionally recalculates formulas
using LibreOffice. Supports both static validation and full recalculation.

Usage:
    python validate_excel.py <excel_file>              # Static validation only
    python validate_excel.py <excel_file> --recalc     # Recalculate with LibreOffice first
    python validate_excel.py <excel_file> --strict     # Also check best practices
"""

import json
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
# Matches cell references including absolute ($A$1, $A1, A$1) and relative (A1)
CELL_REF_PATTERN = re.compile(r"\$?([A-Z]+)\$?(\d+)")
# Matches cross-sheet references: 'Sheet-Name'!A1 or SheetName!A1
# Handles quoted names with special chars (hyphens, spaces, etc.)
CROSS_SHEET_PATTERN = re.compile(r"(?:'([^']+)'|\"([^\"]+)\"|([\w]+))!\$?([A-Z]+)\$?(\d+)")


def recalculate_with_libreoffice(filepath: str, output_dir: str, timeout: int = 60) -> dict:
    """
    Recalculate formulas using LibreOffice WITHOUT modifying the original file.

    Converts xlsx -> ods (which forces recalculation) and saves the ODS file
    to output_dir. The original xlsx is never modified, preserving all formatting.

    Args:
        filepath: Path to the original Excel file
        output_dir: Directory to save the recalculated ODS file
        timeout: Timeout in seconds for LibreOffice operations

    Returns:
        dict with 'success' bool, 'error' message if failed, and 'ods_path' if successful
    """
    # Check if LibreOffice is available
    if not shutil.which("soffice"):
        return {"success": False, "error": "LibreOffice not installed"}

    filepath = Path(filepath)
    if not filepath.exists():
        return {"success": False, "error": f"File not found: {filepath}"}

    # Set environment to avoid display issues
    env = os.environ.copy()
    env["SAL_USE_VCLPLUGIN"] = "svp"  # Use headless VCL plugin
    env["HOME"] = output_dir  # Use output dir for LO config

    # Copy input to output dir first (LibreOffice needs to read from there)
    temp_input = Path(output_dir) / filepath.name
    shutil.copy2(filepath, temp_input)

    # Convert xlsx to ods (forces complete recalculation)
    # We only convert to ODS because openpyxl can read ODS files with ezodf,
    # but actually we'll read it with a different approach - just check for errors
    cmd_to_ods = [
        "soffice",
        "--headless",
        "--norestore",
        "--nofirststartwizard",
        "--convert-to",
        "ods",
        "--outdir",
        output_dir,
        str(temp_input),
    ]

    try:
        result = subprocess.run(
            cmd_to_ods, capture_output=True, text=True, timeout=timeout, env=env
        )

        if result.returncode != 0:
            return {
                "success": False,
                "error": result.stderr or f"xlsx->ods failed with code {result.returncode}",
            }

        # Check for the ODS file
        ods_file = Path(output_dir) / (filepath.stem + ".ods")
        if not ods_file.exists():
            return {"success": False, "error": "ODS conversion failed - no output file"}

        # Now convert ODS to XLSX to get calculated values readable by openpyxl
        # This is a fresh xlsx with calculated values baked in
        cmd_to_xlsx = [
            "soffice",
            "--headless",
            "--norestore",
            "--nofirststartwizard",
            "--convert-to",
            "xlsx",
            "--outdir",
            output_dir,
            str(ods_file),
        ]

        result = subprocess.run(
            cmd_to_xlsx, capture_output=True, text=True, timeout=timeout, env=env
        )

        if result.returncode != 0:
            return {
                "success": False,
                "error": result.stderr or f"ods->xlsx failed with code {result.returncode}",
            }

        # Return path to the recalculated xlsx (NOT the original)
        recalc_xlsx = Path(output_dir) / (filepath.stem + ".xlsx")
        if not recalc_xlsx.exists():
            return {"success": False, "error": "XLSX conversion failed - no output file"}

        return {"success": True, "recalc_path": str(recalc_xlsx)}

    except subprocess.TimeoutExpired:
        return {"success": False, "error": f"Recalculation timed out after {timeout}s"}
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_max_dimensions(ws) -> tuple[int, int]:
    """Get the actual used dimensions of a worksheet."""
    max_row = 0
    max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def find_formula_cells(ws) -> list[dict]:
    """Find all cells containing formulas."""
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(
                    {
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "row": cell.row,
                        "column": cell.column,
                    }
                )
    return formulas


def find_error_cells(ws) -> dict[str, list[str]]:
    """Find cells containing Excel error values.

    Only matches cells whose value IS an error, not cells that contain
    error strings as text (e.g., formulas like =IFERROR(A1/B1,"#N/A")).
    """
    errors = {err: [] for err in EXCEL_ERRORS}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip()
                # Skip formulas - we want calculated values, not formula text
                if val.startswith("="):
                    continue
                # Check if cell value is exactly an error (not a substring match)
                if val in EXCEL_ERRORS:
                    errors[val].append(cell.coordinate)

    return {k: v for k, v in errors.items() if v}


def check_formula_references(formula: str, sheet_names: list[str]) -> list[str]:
    """Check if formula references are valid."""
    issues = []

    # Check cross-sheet references
    # Pattern groups: (1) 'quoted' (2) "quoted" (3) unquoted (4) col (5) row
    for match in CROSS_SHEET_PATTERN.finditer(formula):
        # Get sheet name from whichever group matched
        ref_sheet = match.group(1) or match.group(2) or match.group(3)
        if ref_sheet and ref_sheet not in sheet_names:
            issues.append(f"References non-existent sheet: {ref_sheet}")

    # Check cell references within Excel bounds
    for match in CELL_REF_PATTERN.finditer(formula):
        col_str, row = match.group(1), int(match.group(2))
        if row > 1048576:
            issues.append(f"Row {row} exceeds Excel maximum")
        try:
            col_idx = column_index_from_string(col_str)
            if col_idx > 16384:
                issues.append(f"Column {col_str} exceeds Excel maximum")
        except ValueError:
            issues.append(f"Invalid column reference: {col_str}")

    return issues


def detect_potential_div_zero(formula: str) -> bool:
    """Detect formulas that might cause #DIV/0! errors."""
    if "/" in formula:
        upper = formula.upper()
        if "IFERROR" not in upper and "IF(" not in upper:
            return True
    return False


def detect_hardcoded_numbers(formula: str) -> list[str]:
    """Detect hardcoded numbers in formulas (bad practice)."""
    allowed = {"0", "1", "2", "12", "100", "365", "52", "4"}
    cleaned = CELL_REF_PATTERN.sub("", formula)
    numbers = re.findall(r"\b(\d+(?:\.\d+)?)\b", cleaned)
    return [f"Hardcoded value: {num}" for num in numbers if num not in allowed and float(num) > 1]


def validate_excel(filepath: str, recalc: bool = False, strict: bool = False) -> dict:
    """
    Validate an Excel file WITHOUT modifying it.

    Args:
        filepath: Path to Excel file
        recalc: If True, recalculate formulas with LibreOffice first
        strict: If True, check for best practices violations

    Returns:
        dict with validation results

    Note:
        The original file is NEVER modified. When recalc=True, formulas are
        recalculated in a temporary copy, preserving the original formatting.
    """
    import tempfile

    if not Path(filepath).exists():
        return {"status": "error", "message": f"File not found: {filepath}"}

    result = {
        "status": "success",
        "file": filepath,
        "recalculated": False,
        "sheets": [],
        "summary": {
            "total_sheets": 0,
            "total_formulas": 0,
            "total_errors": 0,
            "total_warnings": 0,
        },
        "errors": {},
        "warnings": [],
    }

    # Path to file with calculated values (original or recalculated copy)
    recalc_file = None
    temp_dir_obj = None

    # Recalculate with LibreOffice if requested
    if recalc:
        temp_dir_obj = tempfile.TemporaryDirectory()
        recalc_result = recalculate_with_libreoffice(filepath, temp_dir_obj.name)
        if recalc_result["success"]:
            result["recalculated"] = True
            recalc_file = recalc_result["recalc_path"]
        else:
            result["warnings"].append(f"Recalculation skipped: {recalc_result['error']}")
            temp_dir_obj.cleanup()
            temp_dir_obj = None

    # Load original workbook (for formulas and structure)
    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        if temp_dir_obj:
            temp_dir_obj.cleanup()
        return {"status": "error", "message": f"Failed to open file: {e}"}

    result["summary"]["total_sheets"] = len(wb.sheetnames)
    sheet_names = wb.sheetnames

    # Load recalculated file to check calculated values for errors
    # This is a SEPARATE file - the original is untouched
    wb_data = None
    if recalc_file:
        try:
            wb_data = load_workbook(recalc_file, data_only=True)
        except Exception:
            result["warnings"].append("Could not read recalculated values")
            wb_data = None

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name] if wb_data else None
        max_row, max_col = get_max_dimensions(ws)

        sheet_info = {
            "name": sheet_name,
            "dimensions": f"{max_row} rows x {max_col} cols",
            "formulas": 0,
            "errors": [],
            "warnings": [],
        }

        # Find formulas
        formulas = find_formula_cells(ws)
        sheet_info["formulas"] = len(formulas)
        result["summary"]["total_formulas"] += len(formulas)

        # Find error values in calculated cells (after recalc)
        if ws_data:
            errors = find_error_cells(ws_data)
        else:
            errors = find_error_cells(ws)

        for err_type, locations in errors.items():
            if err_type not in result["errors"]:
                result["errors"][err_type] = {"count": 0, "locations": []}
            result["errors"][err_type]["count"] += len(locations)
            result["errors"][err_type]["locations"].extend(
                [f"{sheet_name}!{loc}" for loc in locations[:10]]
            )
            result["summary"]["total_errors"] += len(locations)
            sheet_info["errors"].extend([f"{err_type} at {loc}" for loc in locations])

        # Validate formula references
        for formula_info in formulas:
            issues = check_formula_references(formula_info["formula"], sheet_names)
            for issue in issues:
                sheet_info["warnings"].append(f"{formula_info['cell']}: {issue}")

            if detect_potential_div_zero(formula_info["formula"]):
                sheet_info["warnings"].append(
                    f"{formula_info['cell']}: Potential #DIV/0! (no IFERROR protection)"
                )

            if strict:
                for hc in detect_hardcoded_numbers(formula_info["formula"]):
                    sheet_info["warnings"].append(f"{formula_info['cell']}: {hc}")

        result["summary"]["total_warnings"] += len(sheet_info["warnings"])
        result["sheets"].append(sheet_info)

    wb.close()
    if wb_data:
        wb_data.close()

    # Clean up temp directory
    if temp_dir_obj:
        temp_dir_obj.cleanup()

    # Update status
    if result["summary"]["total_errors"] > 0:
        result["status"] = "errors_found"
    elif result["summary"]["total_warnings"] > 0 and strict:
        result["status"] = "warnings_found"

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_excel.py <excel_file> [options]")
        print()
        print("Options:")
        print("  --recalc    Recalculate formulas with LibreOffice first")
        print("  --strict    Also check for best practice violations")
        print()
        print("Examples:")
        print("  python validate_excel.py report.xlsx")
        print("  python validate_excel.py report.xlsx --recalc")
        print("  python validate_excel.py report.xlsx --recalc --strict")
        sys.exit(1)

    filepath = sys.argv[1]
    recalc = "--recalc" in sys.argv
    strict = "--strict" in sys.argv

    result = validate_excel(filepath, recalc=recalc, strict=strict)
    print(json.dumps(result, indent=2))

    if result["status"] == "errors_found":
        sys.exit(1)
    elif result["status"] == "error":
        sys.exit(2)


if __name__ == "__main__":
    main()
