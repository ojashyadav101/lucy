"""File generation tools for Lucy.

Generates PDF, Excel, and CSV files from structured data,
then uploads them to Slack for sharing.

Architecture:
    - PDF: HTML → WeasyPrint (pip install weasyprint)
    - Excel: openpyxl (pip install openpyxl)
    - CSV: stdlib csv module
    - Upload: Slack files_upload_v2 API

All files are generated in a temp directory, uploaded, then cleaned up.
The agent sees these as internal tools (lucy_* prefix).

Dependencies (add to pyproject.toml):
    weasyprint >= 62.0
    openpyxl >= 3.1.0
"""

from __future__ import annotations

import csv
import io
import json
import tempfile
from pathlib import Path
from typing import Any

import structlog

logger = structlog.get_logger()

# ═══════════════════════════════════════════════════════════════════════════
# PDF Generation
# ═══════════════════════════════════════════════════════════════════════════

_PDF_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
<style>
  @page {{ size: A4; margin: 2cm; }}
  body {{ font-family: 'Helvetica Neue', Arial, sans-serif; line-height: 1.6;
         color: #1a1a1a; font-size: 10pt; }}
  h1 {{ color: #1e3a5f; font-size: 20pt; border-bottom: 2px solid #1e3a5f;
       padding-bottom: 6px; }}
  h2 {{ color: #2a5a8f; font-size: 14pt; margin-top: 20px; }}
  h3 {{ color: #333; font-size: 11pt; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0; }}
  th {{ background: #f0f2f5; padding: 8px; text-align: left; font-weight: 600;
       border-bottom: 2px solid #ddd; }}
  td {{ padding: 8px; border-bottom: 1px solid #eee; }}
  code {{ background: #f0f2f5; padding: 2px 6px; border-radius: 3px;
         font-family: monospace; font-size: 9pt; }}
  pre {{ background: #1e1e2e; color: #cdd6f4; padding: 14px;
        border-radius: 6px; font-size: 9pt; overflow-x: auto; }}
  .header {{ text-align: center; margin-bottom: 20px; }}
  .footer {{ text-align: center; margin-top: 30px; color: #888;
            font-size: 8pt; }}
</style>
</head>
<body>
{content}
</body>
</html>"""


async def generate_pdf(
    title: str,
    content_html: str,
    filename: str | None = None,
) -> Path:
    """Generate a PDF from HTML content.

    Args:
        title: Document title (used in filename if none provided).
        content_html: HTML body content (can include h1, h2, tables, etc).
        filename: Output filename. Auto-generated from title if None.

    Returns:
        Path to the generated PDF file.
    """
    try:
        from weasyprint import HTML
    except ImportError:
        raise RuntimeError(
            "PDF generation is not available right now. "
            "Please respond with the content directly in Slack instead."
        )

    full_html = _PDF_TEMPLATE.format(content=content_html)

    if not filename:
        safe_title = "".join(
            c if c.isalnum() or c in " -_" else ""
            for c in title
        ).strip().replace(" ", "_")[:50]
        filename = f"{safe_title}.pdf"

    output_path = Path(tempfile.mkdtemp()) / filename
    try:
        HTML(string=full_html).write_pdf(str(output_path))
    except Exception as e:
        logger.error("pdf_generation_failed", error=str(e))
        raise RuntimeError(
            f"PDF generation failed. Share the content directly in Slack instead. "
            f"Content preview: {content_html[:200]}"
        )

    logger.info(
        "pdf_generated",
        title=title,
        path=str(output_path),
        size_kb=round(output_path.stat().st_size / 1024, 1),
    )
    return output_path


# ═══════════════════════════════════════════════════════════════════════════
# Excel Generation
# ═══════════════════════════════════════════════════════════════════════════

async def generate_excel(
    title: str,
    sheets: dict[str, list[list[Any]]],
    filename: str | None = None,
) -> Path:
    """Generate an Excel file from tabular data.

    Args:
        title: Workbook title (used in filename).
        sheets: Dict of sheet_name → list of rows (first row = headers).
            Example: {"Revenue": [["Month", "MRR"], ["Jan", 18000], ...]}
        filename: Output filename. Auto-generated if None.

    Returns:
        Path to the generated .xlsx file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError(
            "openpyxl not installed. Add to pyproject.toml: "
            "openpyxl >= 3.1.0"
        )

    wb = Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="F0F2F5", end_color="F0F2F5", fill_type="solid",
    )
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for i, (sheet_name, rows) in enumerate(sheets.items()):
        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        for row_idx, row in enumerate(rows):
            if not isinstance(row, (list, tuple)):
                row = [row]
            for col_idx, value in enumerate(row):
                if isinstance(value, (list, tuple, dict)):
                    value = str(value)
                cell = ws.cell(
                    row=row_idx + 1, column=col_idx + 1, value=value,
                )
                if row_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = thin_border

        for col_idx in range(len(rows[0]) if rows else 0):
            max_length = 0
            col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
            for row in rows:
                if col_idx < len(row):
                    max_length = max(max_length, len(str(row[col_idx])))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    if not filename:
        safe_title = "".join(
            c if c.isalnum() or c in " -_" else ""
            for c in title
        ).strip().replace(" ", "_")[:50]
        filename = f"{safe_title}.xlsx"

    output_path = Path(tempfile.mkdtemp()) / filename
    wb.save(str(output_path))

    logger.info(
        "excel_generated",
        title=title,
        sheets=list(sheets.keys()),
        path=str(output_path),
    )
    return output_path


# ═══════════════════════════════════════════════════════════════════════════
# CSV Generation
# ═══════════════════════════════════════════════════════════════════════════

async def generate_csv(
    title: str,
    rows: list[list[Any]],
    filename: str | None = None,
) -> Path:
    """Generate a CSV file from tabular data.

    Args:
        title: Used in filename if none provided.
        rows: List of rows (first row = headers).
        filename: Output filename. Auto-generated if None.

    Returns:
        Path to the generated .csv file.
    """
    if not filename:
        safe_title = "".join(
            c if c.isalnum() or c in " -_" else ""
            for c in title
        ).strip().replace(" ", "_")[:50]
        filename = f"{safe_title}.csv"

    output_path = Path(tempfile.mkdtemp()) / filename

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)

    logger.info("csv_generated", title=title, rows=len(rows))
    return output_path


# ═══════════════════════════════════════════════════════════════════════════
# Slack Upload
# ═══════════════════════════════════════════════════════════════════════════

async def upload_file_to_slack(
    slack_client: Any,
    file_path: Path,
    channel_id: str,
    thread_ts: str | None = None,
    title: str | None = None,
    comment: str | None = None,
) -> dict[str, Any]:
    """Upload a file to Slack and post it to a channel/thread.

    Args:
        slack_client: Slack Bolt async client.
        file_path: Path to the file to upload.
        channel_id: Target channel.
        thread_ts: Thread to post in (optional).
        title: Display title for the file.
        comment: Message text to accompany the file.

    Returns:
        Slack API response dict.
    """
    kwargs: dict[str, Any] = {
        "channel": channel_id,
        "file": str(file_path),
        "filename": file_path.name,
    }
    if title:
        kwargs["title"] = title
    if comment:
        kwargs["initial_comment"] = comment
    if thread_ts:
        kwargs["thread_ts"] = thread_ts

    try:
        result = await slack_client.files_upload_v2(**kwargs)
        logger.info(
            "file_uploaded_to_slack",
            filename=file_path.name,
            channel=channel_id,
        )
        return result
    except Exception as e:
        logger.error(
            "slack_file_upload_failed",
            filename=file_path.name,
            error=str(e),
        )
        try:
            result = await slack_client.files_upload(
                channels=channel_id,
                file=str(file_path),
                filename=file_path.name,
                title=title or file_path.name,
                initial_comment=comment or "",
                thread_ts=thread_ts,
            )
            return result
        except Exception as e2:
            logger.error("slack_file_upload_v1_failed", error=str(e2))
            raise


# ═══════════════════════════════════════════════════════════════════════════
# Tool Definitions (for agent integration)
# ═══════════════════════════════════════════════════════════════════════════

def get_file_tool_definitions() -> list[dict[str, Any]]:
    """Return OpenAI-format tool definitions for file generation and editing.

    Registered as internal tools alongside Slack history search.
    """
    return [
        {
            "type": "function",
            "function": {
                "name": "lucy_write_file",
                "description": (
                    "Write a new file to the workspace. "
                    "Use this to draft a brand new file from scratch. "
                    "If the file already exists, it will be completely overwritten. "
                    "To modify an existing file, use `lucy_edit_file` instead."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Absolute path to the file to create or overwrite.",
                        },
                        "content": {
                            "type": "string",
                            "description": "The complete content to write into the file.",
                        },
                    },
                    "required": ["path", "content"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "lucy_edit_file",
                "description": (
                    "Apply a SEARCH/REPLACE block to edit an existing file. "
                    "This is the ONLY way to fix or modify existing code. "
                    "The `old_string` must match the file content EXACTLY, "
                    "including all whitespace, indentation, and blank lines. "
                    "Provide enough context lines before and after the change "
                    "to uniquely identify the block to replace. "
                    "Do NOT use this to write a new file from scratch."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Absolute path to the file.",
                        },
                        "old_string": {
                            "type": "string",
                            "description": "The exact block of text to replace. Must include enough context to be unique.",
                        },
                        "new_string": {
                            "type": "string",
                            "description": "The new text that will replace the old text.",
                        },
                    },
                    "required": ["path", "old_string", "new_string"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "lucy_store_api_key",
                "description": (
                    "Store a user-provided API key for a custom integration. "
                    "Use this after a user pastes their API key in chat. "
                    "The key is stored securely and used automatically when "
                    "calling tools from that custom integration."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "service_slug": {
                            "type": "string",
                            "description": (
                                "The slug of the custom integration "
                                "(e.g. 'polarsh', 'clerk')."
                            ),
                        },
                        "api_key": {
                            "type": "string",
                            "description": "The API key or token provided by the user.",
                        },
                    },
                    "required": ["service_slug", "api_key"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "lucy_generate_pdf",
                "description": (
                    "Generate a formatted PDF document. ONLY use this when the user "
                    "EXPLICITLY asks for a PDF, report, or document — or when the "
                    "output is genuinely multi-page (5+ sections with data). "
                    "DO NOT use this for short answers, simple questions, lists, "
                    "or anything that fits comfortably in a Slack message. "
                    "When in doubt, respond in text — PDF is for heavy deliverables only. "
                    "Provide HTML content for the body."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title": {
                            "type": "string",
                            "description": "Document title. Appears in filename.",
                        },
                        "content_html": {
                            "type": "string",
                            "description": (
                                "HTML body content. Use h1/h2 for headers, "
                                "table/tr/td for data, p for paragraphs, "
                                "ul/li for lists, code/pre for code."
                            ),
                        },
                    },
                    "required": ["title", "content_html"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "lucy_generate_excel",
                "description": (
                    "Generate an Excel spreadsheet. ONLY use when the user "
                    "explicitly asks for a spreadsheet, Excel file, or .xlsx "
                    "download. Do NOT use for general data display."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title": {
                            "type": "string",
                            "description": "Workbook title. Used in filename.",
                        },
                        "sheets": {
                            "type": "object",
                            "description": (
                                "Object mapping sheet names to row arrays. "
                                "First row in each sheet is the header row. "
                                'Example: {"Revenue": [["Month", "MRR"], '
                                '["Jan", 18000]]}'
                            ),
                            "additionalProperties": {
                                "type": "array",
                                "items": {
                                    "type": "array",
                                    "items": {},
                                },
                            },
                        },
                    },
                    "required": ["title", "sheets"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "lucy_generate_csv",
                "description": (
                    "Generate a CSV file from tabular data. Use for simple "
                    "data exports that don't need Excel formatting."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title": {
                            "type": "string",
                            "description": "Used in filename.",
                        },
                        "rows": {
                            "type": "array",
                            "description": (
                                "Array of row arrays. First row is headers. "
                                'Example: [["Name", "Email"], '
                                '["Alice", "a@b.com"]]'
                            ),
                            "items": {
                                "type": "array",
                                "items": {},
                            },
                        },
                    },
                    "required": ["title", "rows"],
                },
            },
        },
    ]


def _resolve_mangled_path(raw_path: str, ws_root: str) -> Path | None:
    """Recover a valid workspace path when the LLM mangles it.

    Common LLM mistakes:
      - /home/user/src/App.tsx
      - ...worktrees/lucy/rdj/[workspace]/src/App.tsx
      - ...worktrees/lucy/rdj/workspace/src/App.tsx
    """
    import re

    # Extract the relative filename from the path (e.g. "src/App.tsx")
    rel_match = re.search(r"(src/.+\.tsx?)$", raw_path)
    if not rel_match:
        return None
    relative = rel_match.group(1)

    # workspace_root is already the top-level workspaces dir
    spaces_root = Path(ws_root)
    if not spaces_root.exists():
        return None

    newest: Path | None = None
    newest_mtime = 0.0
    for ws_dir in spaces_root.iterdir():
        sp = ws_dir / "spaces"
        if not sp.is_dir():
            continue
        for proj in sp.iterdir():
            if not proj.is_dir():
                continue
            config = proj / "project.json"
            if config.exists():
                mtime = config.stat().st_mtime
                if mtime > newest_mtime:
                    newest_mtime = mtime
                    newest = proj

    if newest is None:
        return None
    return newest / relative


async def execute_file_tool(
    tool_name: str,
    parameters: dict[str, Any],
    slack_client: Any | None = None,
    channel_id: str | None = None,
    thread_ts: str | None = None,
) -> dict[str, Any]:
    """Execute a file generation tool and optionally upload to Slack.

    Returns:
        {"result": "...", "file_path": "/tmp/.../file.pdf"} on success.
        {"error": "..."} on failure.
    """
    try:
        if tool_name == "lucy_write_file":
            path_str = parameters.get("path")
            content = parameters.get("content")
            if not path_str or content is None:
                return {"error": "Missing required parameters: path, content"}

            p = Path(path_str)

            from lucy.config import settings
            ws_root = str(settings.workspace_root)
            resolved = str(p.resolve())

            if not resolved.startswith(ws_root):
                p = _resolve_mangled_path(path_str, ws_root)
                if p is None:
                    logger.warning(
                        "file_write_rejected",
                        original_path=path_str,
                        reason="outside workspace, could not auto-correct",
                    )
                    return {
                        "error": (
                            f"Cannot write to {path_str} — path is invalid. "
                            f"Use the exact app_tsx_path returned by "
                            f"lucy_spaces_init."
                        ),
                    }
                logger.info(
                    "file_write_path_corrected",
                    original=path_str,
                    corrected=str(p),
                )

            logger.info(
                "file_write_attempt",
                path=str(p),
                content_length=len(content) if content else 0,
            )

            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(content, encoding="utf-8")
            logger.info(
                "file_write_success",
                path=str(p),
                chars=len(content),
            )
            return {
                "result": f"Successfully wrote {len(content)} characters to {p}",
                "file_path": str(p),
            }

        elif tool_name == "lucy_edit_file":
            path_str = parameters.get("path")
            old_string = parameters.get("old_string")
            new_string = parameters.get("new_string")
            if not path_str or old_string is None or new_string is None:
                return {"error": "Missing required parameters: path, old_string, new_string"}
            
            p = Path(path_str)
            if not p.exists():
                return {"error": f"File not found: {path_str}"}
                
            file_content = p.read_text(encoding="utf-8")
            if old_string not in file_content:
                return {
                    "error": (
                        "old_string did not match exactly. "
                        "Make sure you include the exact whitespace, indentation, "
                        "and context lines. Do not omit any lines in the middle "
                        "of the block."
                    )
                }
                
            if file_content.count(old_string) > 1:
                return {
                    "error": (
                        "old_string matched multiple times. "
                        "Please provide more context lines before or after "
                        "to make the replacement block unique."
                    )
                }
                
            new_content = file_content.replace(old_string, new_string)
            p.write_text(new_content, encoding="utf-8")
            return {
                "result": f"Successfully edited {path_str} (replaced {len(old_string)} chars with {len(new_string)} chars)",
                "file_path": str(path_str),
            }
            
        elif tool_name == "lucy_generate_pdf":
            content_html = parameters.get("content_html", "")
            if len(content_html.strip()) < 200:
                return {
                    "error": (
                        "This content is too short for a PDF. "
                        "Respond directly in Slack instead. "
                        "PDFs should be used for multi-page reports "
                        "and detailed documents only."
                    ),
                }
            path = await generate_pdf(
                title=parameters["title"],
                content_html=content_html,
            )
        elif tool_name == "lucy_generate_excel":
            sheets = parameters.get("sheets", {})
            if isinstance(sheets, str):
                sheets = json.loads(sheets)
            if isinstance(sheets, list):
                converted: dict[str, list[list[Any]]] = {}
                for i, item in enumerate(sheets):
                    if isinstance(item, dict):
                        name = item.get("name", item.get("sheet_name", f"Sheet{i+1}"))
                        rows = item.get("rows", item.get("data", []))
                        converted[name] = rows
                    elif isinstance(item, list):
                        converted[f"Sheet{i+1}"] = item
                sheets = converted
            path = await generate_excel(
                title=parameters["title"],
                sheets=sheets,
            )
        elif tool_name == "lucy_generate_csv":
            rows = parameters.get("rows", [])
            if isinstance(rows, str):
                rows = json.loads(rows)
            path = await generate_csv(
                title=parameters["title"],
                rows=rows,
            )
        else:
            return {"error": f"Unknown file tool: {tool_name}"}

        upload_result = None
        if slack_client and channel_id:
            upload_result = await upload_file_to_slack(
                slack_client=slack_client,
                file_path=path,
                channel_id=channel_id,
                thread_ts=thread_ts,
                title=parameters.get("title", path.name),
            )

        result = {
            "result": f"Generated {path.name} ({path.stat().st_size} bytes)",
            "file_path": str(path),
            "filename": path.name,
        }
        if upload_result:
            result["uploaded"] = True
            result["result"] += " — uploaded to Slack."

        return result

    except Exception as e:
        logger.error("file_tool_failed", tool=tool_name, error=str(e))
        return {"error": str(e)}
