"""
A Python wrapper module for the Polar.sh API, designed for use by an AI assistant.

This module provides a set of tools to interact with the Polar.sh API, covering
various aspects like product management, subscriptions, customer management,
checkout flows, orders, and webhooks. It uses httpx for asynchronous HTTP
requests and handles authentication via a Bearer token.
"""

from __future__ import annotations

import httpx
import json
import tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from lucy.config import settings

API_BASE_URL = "https://api.polar.sh"

async def _make_request(
    method: str,
    url: str,
    api_key: str,
    params: dict = None,
    json_data: dict = None,
) -> dict:
    headers = {"Authorization": f"Bearer {api_key}"}
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=settings.polar_api_timeout_s) as client:
            response = await client.request(
                method, url, headers=headers, params=params, json=json_data
            )
            response.raise_for_status()
            return response.json()
    except httpx.HTTPStatusError as e:
        return {"error": f"API error: {e.response.status_code} - {e.response.text}"}
    except httpx.RequestError as e:
        return {"error": f"Request error: {e}"}
    except json.JSONDecodeError:
        return {"error": f"Invalid JSON response: {response.text}"}
    except Exception as e:
        return {"error": f"An unexpected error occurred: {e}"}


TOOLS = [
    {
        "name": "polarsh_list_products",
        "description": "List products available on Polar.sh. Supports filtering by ID, organization ID, query, archived status, and recurring status. Use `page` and `limit` for pagination.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "Product ID to filter by."},
                "organization_id": {"type": "string", "description": "Organization ID to filter by."},
                "query": {"type": "string", "description": "Search query for product names or descriptions."},
                "is_archived": {"type": "boolean", "description": "Filter by archived status."},
                "is_recurring": {"type": "boolean", "description": "Filter by recurring products."},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
                "limit": {"type": "integer", "description": "Number of items per page.", "default": 10},
            },
        },
    },
    {
        "name": "polarsh_create_product",
        "description": "Create a new product on Polar.sh.",
        "parameters": {
            "type": "object",
            "properties": {
                "name": {"type": "string", "description": "Name of the product.", "min_length": 1},
                "organization_id": {"type": "string", "description": "ID of the organization that owns the product.", "min_length": 1},
                "type": {"type": "string", "description": "Type of the product (e.g., 'product', 'subscription').", "enum": ["product", "subscription"], "min_length": 1},
                "description": {"type": "string", "description": "Description of the product."},
                "prices": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "price_amount": {"type": "integer", "description": "Amount in cents."},
                            "price_currency": {"type": "string", "description": "Currency code (e.g., 'USD')."},
                            "recurring_interval": {"type": "string", "description": "Interval for recurring prices (e.g., 'month', 'year')."},
                            "recurring_interval_count": {"type": "integer", "description": "Number of intervals for recurring prices."},
                        },
                        "required": ["price_amount", "price_currency"],
                    },
                    "description": "List of prices for the product.",
                },
            },
            "required": ["name", "organization_id", "type"],
        },
    },
    {
        "name": "polarsh_get_product",
        "description": "Retrieve a specific product by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the product to retrieve.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_update_product",
        "description": "Update an existing product on Polar.sh.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "The ID of the product to update.", "min_length": 1},
                "name": {"type": "string", "description": "New name of the product."},
                "description": {"type": "string", "description": "New description of the product."},
                "is_archived": {"type": "boolean", "description": "Whether the product is archived."},
            },
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_list_subscriptions",
        "description": "List subscriptions. Supports filtering by organization ID, product ID, customer ID, external customer ID, and discount ID. Use `page` and `limit` for pagination.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "Organization ID to filter by."},
                "product_id": {"type": "string", "description": "Product ID to filter by."},
                "customer_id": {"type": "string", "description": "Customer ID to filter by."},
                "external_customer_id": {"type": "string", "description": "External Customer ID to filter by."},
                "discount_id": {"type": "string", "description": "Discount ID to filter by."},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
                "limit": {"type": "integer", "description": "Number of items per page.", "default": 10},
            },
        },
    },
    {
        "name": "polarsh_create_subscription",
        "description": "Create a new subscription for a customer.",
        "parameters": {
            "type": "object",
            "properties": {
                "product_id": {"type": "string", "description": "ID of the product to subscribe to.", "min_length": 1},
                "customer_id": {"type": "string", "description": "ID of the customer.", "min_length": 1},
                "price_id": {"type": "string", "description": "ID of the price to use for the subscription.", "min_length": 1},
                "starts_at": {"type": "string", "format": "date-time", "description": "Timestamp when the subscription starts."},
                "cancel_at_period_end": {"type": "boolean", "description": "Whether to cancel the subscription at the end of the current period."},
            },
            "required": ["product_id", "customer_id", "price_id"],
        },
    },
    {
        "name": "polarsh_get_subscription",
        "description": "Retrieve a specific subscription by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the subscription to retrieve.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_update_subscription",
        "description": "Update an existing subscription.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "The ID of the subscription to update.", "min_length": 1},
                "cancel_at_period_end": {"type": "boolean", "description": "Whether to cancel the subscription at the end of the current period."},
                "price_id": {"type": "string", "description": "New price ID for the subscription."},
            },
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_revoke_subscription",
        "description": "Revoke (cancel) a subscription by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the subscription to revoke.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_list_customers",
        "description": "List Polar.sh customers. For Excel reports: pass export_excel=true to generate a formatted Excel file with ALL customers. Returns a file path you can upload to Slack. For quick lookups: use default pagination.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "Organization ID to filter by."},
                "email": {"type": "string", "description": "Customer email to filter by."},
                "query": {"type": "string", "description": "Search query for customer names or emails."},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
                "limit": {"type": "integer", "description": "Number of items per page.", "default": 10},
                "export_excel": {"type": "string", "description": "Set to 'true' to generate a formatted Excel file with ALL customers.", "default": "false"},
                "return_all": {"type": "string", "description": "Set to 'true' to return ALL customers as JSON.", "default": "false"},
            },
        },
    },
    {
        "name": "polarsh_create_customer",
        "description": "Create a new customer.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "ID of the organization the customer belongs to.", "min_length": 1},
                "email": {"type": "string", "format": "email", "description": "Email address of the customer.", "min_length": 1},
                "name": {"type": "string", "description": "Name of the customer."},
                "external_id": {"type": "string", "description": "An optional external ID for the customer."},
            },
            "required": ["organization_id", "email"],
        },
    },
    {
        "name": "polarsh_get_customer",
        "description": "Retrieve a specific customer by their ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the customer to retrieve.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_update_customer",
        "description": "Update an existing customer.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "The ID of the customer to update.", "min_length": 1},
                "email": {"type": "string", "format": "email", "description": "New email address of the customer."},
                "name": {"type": "string", "description": "New name of the customer."},
            },
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_delete_customer",
        "description": "Delete a customer by their ID. Can optionally anonymize the customer data.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "The ID of the customer to delete.", "min_length": 1},
                "anonymize": {"type": "boolean", "description": "Whether to anonymize the customer data instead of full deletion.", "default": False},
            },
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_create_checkout_link",
        "description": "Create a new checkout link for a product.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "ID of the organization.", "min_length": 1},
                "product_id": {"type": "string", "description": "ID of the product for the checkout link.", "min_length": 1},
                "price_id": {"type": "string", "description": "ID of the price to use for the checkout.", "min_length": 1},
                "success_url": {"type": "string", "format": "uri", "description": "URL to redirect to after successful checkout."},
                "cancel_url": {"type": "string", "format": "uri", "description": "URL to redirect to if checkout is cancelled."},
                "customer_id": {"type": "string", "description": "Optional customer ID to pre-fill checkout."},
            },
            "required": ["organization_id", "product_id", "price_id", "success_url", "cancel_url"],
        },
    },
    {
        "name": "polarsh_list_orders",
        "description": "List Polar.sh orders. For Excel reports: pass export_excel=true to generate a formatted Excel file with ALL orders. Returns a file path you can upload to Slack.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "Organization ID to filter by."},
                "product_id": {"type": "string", "description": "Product ID to filter by."},
                "product_billing_type": {"type": "string", "description": "Billing type of the product (e.g., 'one_time', 'recurring')."},
                "discount_id": {"type": "string", "description": "Discount ID to filter by."},
                "customer_id": {"type": "string", "description": "Customer ID to filter by."},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
                "limit": {"type": "integer", "description": "Number of items per page.", "default": 10},
                "export_excel": {"type": "string", "description": "Set to 'true' to generate a formatted Excel file with ALL orders.", "default": "false"},
                "return_all": {"type": "string", "description": "Set to 'true' to return ALL orders as JSON.", "default": "false"},
            },
        },
    },
    {
        "name": "polarsh_get_order",
        "description": "Retrieve a specific order by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the order to retrieve.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_generate_order_invoice",
        "action_type": "READ",
        "description": "Generate an invoice for a specific order.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the order to generate an invoice for.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_list_webhooks_endpoints",
        "description": "List webhook endpoints for an organization. Use `page` and `limit` for pagination.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "Organization ID to filter by.", "min_length": 1},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
                "limit": {"type": "integer", "description": "Number of items per page.", "default": 10},
            },
            "required": ["organization_id"],
        },
    },
    {
        "name": "polarsh_create_webhook_endpoint",
        "description": "Create a new webhook endpoint.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "ID of the organization.", "min_length": 1},
                "url": {"type": "string", "format": "uri", "description": "The URL where webhook events will be sent.", "min_length": 1},
                "secret": {"type": "string", "description": "Optional secret for signing webhook payloads."},
                "enabled_events": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of event types to send to this endpoint (e.g., 'order.created', 'subscription.updated').",
                },
            },
            "required": ["organization_id", "url"],
        },
    },
    {
        "name": "polarsh_get_webhook_endpoint",
        "description": "Retrieve a specific webhook endpoint by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the webhook endpoint to retrieve.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_update_webhook_endpoint",
        "description": "Update an existing webhook endpoint.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "The ID of the webhook endpoint to update.", "min_length": 1},
                "url": {"type": "string", "format": "uri", "description": "New URL for the webhook endpoint."},
                "enabled_events": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "New list of event types to send to this endpoint.",
                },
                "disabled": {"type": "boolean", "description": "Whether the webhook endpoint is disabled."},
            },
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_delete_webhook_endpoint",
        "description": "Delete a webhook endpoint by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the webhook endpoint to delete.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_list_benefit_grants",
        "description": "List benefit grants. Supports filtering by organization ID, customer ID, external customer ID, and granted status. Use `page` for pagination.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "Organization ID to filter by."},
                "customer_id": {"type": "string", "description": "Customer ID to filter by."},
                "external_customer_id": {"type": "string", "description": "External Customer ID to filter by."},
                "is_granted": {"type": "boolean", "description": "Filter by whether the benefit is granted."},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
            },
        },
    },
    {
        "name": "polarsh_list_benefits",
        "description": "List benefits. Supports filtering by organization ID, type, ID, excluding ID, and query. Use `page` and `limit` for pagination.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "Organization ID to filter by."},
                "type": {"type": "string", "description": "Benefit type to filter by."},
                "id": {"type": "string", "description": "Benefit ID to filter by."},
                "exclude_id": {"type": "string", "description": "Benefit ID to exclude from results."},
                "query": {"type": "string", "description": "Search query for benefit names or descriptions."},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
                "limit": {"type": "integer", "description": "Number of items per page.", "default": 10},
            },
        },
    },
    {
        "name": "polarsh_create_benefit",
        "description": "Create a new benefit.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "ID of the organization.", "min_length": 1},
                "type": {"type": "string", "description": "Type of the benefit (e.g., 'ads', 'discord').", "min_length": 1},
                "description": {"type": "string", "description": "Description of the benefit.", "min_length": 1},
                "is_tax_applicable": {"type": "boolean", "description": "Whether the benefit is tax applicable."},
                "selectable": {"type": "boolean", "description": "Whether the benefit is selectable by customers."},
                "properties": {"type": "object", "description": "Additional properties for the benefit, depends on type."},
            },
            "required": ["organization_id", "type", "description"],
        },
    },
    {
        "name": "polarsh_get_benefit",
        "description": "Retrieve a specific benefit by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the benefit to retrieve.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_update_benefit",
        "description": "Update an existing benefit.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "The ID of the benefit to update.", "min_length": 1},
                "description": {"type": "string", "description": "New description of the benefit."},
                "is_tax_applicable": {"type": "boolean", "description": "Whether the benefit is tax applicable."},
                "selectable": {"type": "boolean", "description": "Whether the benefit is selectable by customers."},
                "properties": {"type": "object", "description": "New additional properties for the benefit."},
            },
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_delete_benefit",
        "description": "Delete a benefit by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the benefit to delete.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_list_discounts",
        "description": "List discounts. Supports filtering by organization ID and query. Use `page`, `limit`, and `sorting` for pagination and ordering.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "Organization ID to filter by."},
                "query": {"type": "string", "description": "Search query for discount codes or names."},
                "page": {"type": "integer", "description": "Page number for pagination.", "default": 1},
                "limit": {"type": "integer", "description": "Number of items per page.", "default": 10},
                "sorting": {"type": "string", "description": "Sorting order (e.g., 'created_at', '-created_at')."},
            },
        },
    },
    {
        "name": "polarsh_create_discount",
        "description": "Create a new discount.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "ID of the organization.", "min_length": 1},
                "code": {"type": "string", "description": "The discount code.", "min_length": 1},
                "type": {"type": "string", "description": "Type of discount (e.g., 'fixed_amount', 'percentage').", "enum": ["fixed_amount", "percentage"], "min_length": 1},
                "value": {"type": "integer", "description": "The discount value (e.g., amount in cents for fixed, percentage for percentage).", "minimum": 0},
                "product_id": {"type": "string", "description": "Optional product ID to apply the discount to."},
                "starts_at": {"type": "string", "format": "date-time", "description": "Timestamp when the discount becomes active."},
                "expires_at": {"type": "string", "format": "date-time", "description": "Timestamp when the discount expires."},
                "usage_limit": {"type": "integer", "description": "Maximum number of times the discount can be used."},
            },
            "required": ["organization_id", "code", "type", "value"],
        },
    },
    {
        "name": "polarsh_get_discount",
        "description": "Retrieve a specific discount by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the discount to retrieve.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_update_discount",
        "description": "Update an existing discount.",
        "parameters": {
            "type": "object",
            "properties": {
                "id": {"type": "string", "description": "The ID of the discount to update.", "min_length": 1},
                "code": {"type": "string", "description": "New discount code."},
                "value": {"type": "integer", "description": "New discount value."},
                "expires_at": {"type": "string", "format": "date-time", "description": "New expiration timestamp."},
                "usage_limit": {"type": "integer", "description": "New usage limit."},
            },
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_delete_discount",
        "description": "Delete a discount by its ID.",
        "parameters": {
            "type": "object",
            "properties": {"id": {"type": "string", "description": "The ID of the discount to delete.", "min_length": 1}},
            "required": ["id"],
        },
    },
    {
        "name": "polarsh_get_metrics",
        "action_type": "READ",
        "description": "Retrieve metrics for an organization within a specified date range and interval.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "ID of the organization to retrieve metrics for.", "min_length": 1},
                "start_date": {"type": "string", "format": "date", "description": "Start date for the metrics (YYYY-MM-DD).", "min_length": 1},
                "end_date": {"type": "string", "format": "date", "description": "End date for the metrics (YYYY-MM-DD).", "min_length": 1},
                "interval": {"type": "string", "description": "Interval for the metrics (e.g., 'day', 'week', 'month').", "enum": ["day", "week", "month"], "min_length": 1},
                "timezone": {"type": "string", "description": "Timezone for the metrics (e.g., 'UTC', 'America/New_York').", "default": "UTC"},
            },
            "required": ["organization_id", "start_date", "end_date", "interval"],
        },
    },
]


async def _paginate_all(
    endpoint: str, api_key: str, params: dict | None = None,
    max_pages: int = 200,
) -> list[dict]:
    """Fetch all records from a paginated Polar.sh endpoint."""
    params = dict(params or {})
    params["limit"] = 100
    params["page"] = 1
    all_items: list[dict] = []
    while params["page"] <= max_pages:
        result = await _make_request("GET", endpoint, api_key, params=params)
        if isinstance(result, dict) and result.get("error"):
            if not all_items:
                return result  # type: ignore[return-value]
            break
        if not isinstance(result, dict):
            break
        items = result.get("items", result.get("result", []))
        if isinstance(items, list):
            all_items.extend(items)
        pagination = result.get("pagination", {})
        max_page = pagination.get("max_page", 1)
        if params["page"] >= max_page:
            break
        params["page"] += 1
    return all_items


def _build_customers_excel(customers: list[dict]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "All Customers"

    headers = ["ID", "Email", "Name", "Created At", "Organization ID"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, cust in enumerate(customers, 2):
        ws.cell(row=r, column=1, value=cust.get("id", ""))
        ws.cell(row=r, column=2, value=cust.get("email", ""))
        ws.cell(row=r, column=3, value=cust.get("name", ""))
        created = cust.get("created_at", "")
        ws.cell(row=r, column=4, value=str(created)[:19] if created else "")
        ws.cell(row=r, column=5, value=cust.get("organization_id", ""))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 25

    summary = wb.create_sheet("Summary")
    summary.cell(row=1, column=1, value="Total Customers").font = Font(bold=True)
    summary.cell(row=1, column=2, value=len(customers))

    tmp_dir = Path(tempfile.mkdtemp())
    file_path = tmp_dir / "Polarsh_Customers_Export.xlsx"
    wb.save(str(file_path))
    return file_path


def _build_orders_excel(orders: list[dict]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "All Orders"

    headers = ["ID", "Customer Email", "Product", "Amount", "Currency", "Status", "Created At"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, order in enumerate(orders, 2):
        ws.cell(row=r, column=1, value=order.get("id", ""))
        cust = order.get("customer", {}) or {}
        ws.cell(row=r, column=2, value=cust.get("email", ""))
        prod = order.get("product", {}) or {}
        ws.cell(row=r, column=3, value=prod.get("name", ""))
        ws.cell(row=r, column=4, value=order.get("amount", 0) / 100)
        ws.cell(row=r, column=5, value=order.get("currency", "USD"))
        ws.cell(row=r, column=6, value=order.get("status", ""))
        created = order.get("created_at", "")
        ws.cell(row=r, column=7, value=str(created)[:19] if created else "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 22

    summary = wb.create_sheet("Summary")
    summary.cell(row=1, column=1, value="Total Orders").font = Font(bold=True)
    summary.cell(row=1, column=2, value=len(orders))
    total_revenue = sum((o.get("amount", 0) for o in orders), 0) / 100
    summary.cell(row=2, column=1, value="Total Revenue").font = Font(bold=True)
    summary.cell(row=2, column=2, value=f"${total_revenue:,.2f}")
    status_counts = Counter(o.get("status", "unknown") for o in orders)
    summary.cell(row=4, column=1, value="By Status").font = Font(bold=True)
    for i, (status, count) in enumerate(status_counts.most_common(), 5):
        summary.cell(row=i, column=1, value=status)
        summary.cell(row=i, column=2, value=count)

    tmp_dir = Path(tempfile.mkdtemp())
    file_path = tmp_dir / "Polarsh_Orders_Export.xlsx"
    wb.save(str(file_path))
    return file_path


async def execute(tool_name: str, args: dict, api_key: str) -> dict:
    if tool_name == "polarsh_list_products":
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/products/", api_key, params=args
        )
    elif tool_name == "polarsh_create_product":
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/products/", api_key, json_data=args
        )
    elif tool_name == "polarsh_get_product":
        product_id = args.pop("id")
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/products/{product_id}", api_key
        )
    elif tool_name == "polarsh_update_product":
        product_id = args.pop("id")
        return await _make_request(
            "PATCH", f"{API_BASE_URL}/v1/products/{product_id}", api_key, json_data=args
        )
    elif tool_name == "polarsh_list_subscriptions":
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/subscriptions/", api_key, params=args
        )
    elif tool_name == "polarsh_create_subscription":
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/subscriptions/", api_key, json_data=args
        )
    elif tool_name == "polarsh_get_subscription":
        subscription_id = args.pop("id")
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/subscriptions/{subscription_id}", api_key
        )
    elif tool_name == "polarsh_update_subscription":
        subscription_id = args.pop("id")
        return await _make_request(
            "PATCH", f"{API_BASE_URL}/v1/subscriptions/{subscription_id}", api_key, json_data=args
        )
    elif tool_name == "polarsh_revoke_subscription":
        subscription_id = args.pop("id")
        return await _make_request(
            "DELETE", f"{API_BASE_URL}/v1/subscriptions/{subscription_id}", api_key
        )
    elif tool_name == "polarsh_list_customers":
        export_excel = str(args.pop("export_excel", "false")).lower() == "true"
        return_all = str(args.pop("return_all", "false")).lower() == "true"
        if export_excel or return_all:
            all_customers = await _paginate_all(
                f"{API_BASE_URL}/v1/customers/", api_key, params=args,
            )
            if export_excel:
                file_path = _build_customers_excel(all_customers)
                return {
                    "total_count": len(all_customers),
                    "excel_file_path": str(file_path),
                    "sheets": ["All Customers", "Summary"],
                    "note": f"Excel created with ALL {len(all_customers)} customers. Upload it to Slack.",
                }
            return {"total_count": len(all_customers), "items": all_customers}
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/customers/", api_key, params=args
        )
    elif tool_name == "polarsh_create_customer":
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/customers/", api_key, json_data=args
        )
    elif tool_name == "polarsh_get_customer":
        customer_id = args.pop("id")
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/customers/{customer_id}", api_key
        )
    elif tool_name == "polarsh_update_customer":
        customer_id = args.pop("id")
        return await _make_request(
            "PATCH", f"{API_BASE_URL}/v1/customers/{customer_id}", api_key, json_data=args
        )
    elif tool_name == "polarsh_delete_customer":
        customer_id = args.pop("id")
        params = {"anonymize": args.get("anonymize", False)}
        return await _make_request(
            "DELETE", f"{API_BASE_URL}/v1/customers/{customer_id}", api_key, params=params
        )
    elif tool_name == "polarsh_create_checkout_link":
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/checkout-links/", api_key, json_data=args
        )
    elif tool_name == "polarsh_list_orders":
        export_excel = str(args.pop("export_excel", "false")).lower() == "true"
        return_all = str(args.pop("return_all", "false")).lower() == "true"
        if export_excel or return_all:
            all_orders = await _paginate_all(
                f"{API_BASE_URL}/v1/orders/", api_key, params=args,
            )
            if export_excel:
                file_path = _build_orders_excel(all_orders)
                return {
                    "total_count": len(all_orders),
                    "excel_file_path": str(file_path),
                    "sheets": ["All Orders", "Summary"],
                    "note": f"Excel created with ALL {len(all_orders)} orders. Upload it to Slack.",
                }
            return {"total_count": len(all_orders), "items": all_orders}
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/orders/", api_key, params=args
        )
    elif tool_name == "polarsh_get_order":
        order_id = args.pop("id")
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/orders/{order_id}", api_key
        )
    elif tool_name == "polarsh_generate_order_invoice":
        order_id = args.pop("id")
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/orders/{order_id}/invoice", api_key
        )
    elif tool_name == "polarsh_list_webhooks_endpoints":
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/webhooks/endpoints", api_key, params=args
        )
    elif tool_name == "polarsh_create_webhook_endpoint":
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/webhooks/endpoints", api_key, json_data=args
        )
    elif tool_name == "polarsh_get_webhook_endpoint":
        endpoint_id = args.pop("id")
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/webhooks/endpoints/{endpoint_id}", api_key
        )
    elif tool_name == "polarsh_update_webhook_endpoint":
        endpoint_id = args.pop("id")
        return await _make_request(
            "PATCH", f"{API_BASE_URL}/v1/webhooks/endpoints/{endpoint_id}", api_key, json_data=args
        )
    elif tool_name == "polarsh_delete_webhook_endpoint":
        endpoint_id = args.pop("id")
        return await _make_request(
            "DELETE", f"{API_BASE_URL}/v1/webhooks/endpoints/{endpoint_id}", api_key
        )
    elif tool_name == "polarsh_list_benefit_grants":
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/benefit-grants/", api_key, params=args
        )
    elif tool_name == "polarsh_list_benefits":
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/benefits/", api_key, params=args
        )
    elif tool_name == "polarsh_create_benefit":
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/benefits/", api_key, json_data=args
        )
    elif tool_name == "polarsh_get_benefit":
        benefit_id = args.pop("id")
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/benefits/{benefit_id}", api_key
        )
    elif tool_name == "polarsh_update_benefit":
        benefit_id = args.pop("id")
        return await _make_request(
            "PATCH", f"{API_BASE_URL}/v1/benefits/{benefit_id}", api_key, json_data=args
        )
    elif tool_name == "polarsh_delete_benefit":
        benefit_id = args.pop("id")
        return await _make_request(
            "DELETE", f"{API_BASE_URL}/v1/benefits/{benefit_id}", api_key
        )
    elif tool_name == "polarsh_list_discounts":
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/discounts/", api_key, params=args
        )
    elif tool_name == "polarsh_create_discount":
        return await _make_request(
            "POST", f"{API_BASE_URL}/v1/discounts/", api_key, json_data=args
        )
    elif tool_name == "polarsh_get_discount":
        discount_id = args.pop("id")
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/discounts/{discount_id}", api_key
        )
    elif tool_name == "polarsh_update_discount":
        discount_id = args.pop("id")
        return await _make_request(
            "PATCH", f"{API_BASE_URL}/v1/discounts/{discount_id}", api_key, json_data=args
        )
    elif tool_name == "polarsh_delete_discount":
        discount_id = args.pop("id")
        return await _make_request(
            "DELETE", f"{API_BASE_URL}/v1/discounts/{discount_id}", api_key
        )
    elif tool_name == "polarsh_get_metrics":
        return await _make_request(
            "GET", f"{API_BASE_URL}/v1/metrics/", api_key, params=args
        )
    else:
        return {"error": f"Tool {tool_name} not found."}