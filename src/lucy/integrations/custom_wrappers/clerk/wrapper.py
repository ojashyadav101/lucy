"""
A production-ready Python wrapper module for the Clerk API, designed for use by an AI assistant.

This wrapper provides a comprehensive set of tools covering major business categories
within Clerk, including user management, organization management, and webhook management.
It uses httpx for asynchronous HTTP requests and handles authentication via a bearer token.

The module defines a `TOOLS` list, which describes the available API functions
for an AI agent, and an `execute` function to dispatch calls to these functions.
"""

import json
import tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import httpx
import structlog
from typing import Any, Dict, List

from lucy.config import settings

logger = structlog.get_logger()

BASE_URL = "https://api.clerk.com/v1"

async def _make_request(
    method: str,
    endpoint: str,
    api_key: str,
    json_data: Dict = None,
    params: Dict = None,
) -> Dict:
    """Helper function to make authenticated HTTP requests to the Clerk API."""
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    url = f"{BASE_URL}/{endpoint}"

    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=settings.clerk_api_timeout_s) as client:
            if method == "GET":
                response = await client.get(url, headers=headers, params=params)
            elif method == "POST":
                response = await client.post(url, headers=headers, json=json_data)
            elif method == "PATCH":
                response = await client.patch(url, headers=headers, json=json_data)
            elif method == "PUT":
                response = await client.put(url, headers=headers, json=json_data)
            elif method == "DELETE":
                response = await client.delete(url, headers=headers)
            else:
                return {"error": f"Unsupported HTTP method: {method}"}

            response.raise_for_status()  # Raise an exception for 4xx or 5xx status codes
            return response.json()
    except httpx.HTTPStatusError as e:
        return {"error": f"HTTP error occurred: {e.response.status_code} - {e.response.text}"}
    except httpx.RequestError as e:
        return {"error": f"An error occurred while requesting {e.request.url!r}: {e}"}
    except json.JSONDecodeError:
        return {"error": f"Failed to decode JSON from response: {response.text}"}
    except Exception as e:
        return {"error": f"An unexpected error occurred: {e}"}

# --- User Management ---

async def clerk_list_users(api_key: str, query_params: Dict = None) -> Dict:
    """
    Lists all users in your Clerk instance. Supports filtering and pagination.
    Auto-paginates to get total count when no specific filters are applied.
    Supports created_after_unix_ms and created_before_unix_ms for date range filtering.
    """
    params = dict(query_params or {})

    def _ensure_ms(val: int | str) -> int:
        """Convert to milliseconds. LLMs sometimes pass seconds instead."""
        ts = int(val)
        if ts < 1e12:
            ts = ts * 1000
        return ts

    if "created_after_unix_ms" in params:
        params["created_at_after"] = _ensure_ms(params.pop("created_after_unix_ms"))
    if "created_before_unix_ms" in params:
        params["created_at_before"] = _ensure_ms(params.pop("created_before_unix_ms"))

    limit = int(params.get("limit", 100))
    offset = int(params.get("offset", 0))
    params["limit"] = min(limit, 500)
    params["offset"] = offset

    has_date_filter = "created_at_after" in params or "created_at_before" in params

    first_page = await _make_request("GET", "users", api_key, params=params)

    if isinstance(first_page, dict) and first_page.get("error"):
        return first_page

    if isinstance(first_page, list) and len(first_page) >= limit and offset == 0:
        all_users = list(first_page)
        current_offset = limit
        while True:
            params["offset"] = current_offset
            params["limit"] = 500
            page = await _make_request("GET", "users", api_key, params=params)
            if not isinstance(page, list) or len(page) == 0:
                break
            all_users.extend(page)
            if len(page) < 500:
                break
            current_offset += 500

        total = len(all_users)
        export_excel = str(params.get("export_excel", "false")).lower() == "true"
        return_all = str(params.get("return_all", "false")).lower() == "true"

        if export_excel:
            file_path = _build_users_excel(all_users, has_date_filter)
            return {
                "total_count": total,
                "excel_file_path": str(file_path),
                "sheets": ["All Users", "Summary"],
                "note": (
                    f"Excel created with ALL {total} users"
                    + (" in the specified time range" if has_date_filter else "")
                    + f". File ready at {file_path}. "
                    + "Upload it to Slack with the file upload tool, "
                    + "or tell the user it's ready."
                ),
            }

        if return_all:
            return {
                "total_count": total,
                "users_shown": _clean_user_list(all_users),
                "note": (
                    f"Returning all {total} users"
                    + (" in the specified time range" if has_date_filter else "")
                    + "."
                ),
            }
        sample_size = min(20, total)
        return {
            "total_count": total,
            "users_shown": _clean_user_list(all_users[:sample_size]),
            "note": (
                f"Found {total} users"
                + (" in the specified time range" if has_date_filter else "")
                + f". Showing a SAMPLE of {sample_size}. "
                + "Pass export_excel=true to create an Excel file with ALL "
                + f"{total} users (2 sheets: raw data + summary). "
                + "Or pass return_all=true for raw JSON data."
            ),
        }

    if isinstance(first_page, list):
        count = len(first_page)
        result = {
            "total_count": count,
            "users_shown": _clean_user_list(first_page),
        }
        if has_date_filter:
            result["note"] = f"Found {count} users in the specified time range."
        return result
    return first_page


def _clean_user_list(users: list) -> list:
    """Return compact user summaries to avoid context window bloat.

    Full Clerk user objects contain 30+ fields and nested email objects.
    The LLM only needs name, email, signup date, and auth method to
    answer the vast majority of user queries.
    """
    compact = []
    for user in users:
        if not isinstance(user, dict):
            compact.append(user)
            continue
        first = (user.get("first_name") or "").strip()
        last = (user.get("last_name") or "").strip()
        name = f"{first} {last}".strip()
        primary_email = ""
        emails = user.get("email_addresses", [])
        if emails and isinstance(emails, list):
            e = emails[0]
            if isinstance(e, dict):
                primary_email = e.get("email_address", "")
            elif isinstance(e, str):
                primary_email = e
        if not name:
            name = primary_email or user.get("username") or "Unknown User"

        created = user.get("created_at")
        if isinstance(created, (int, float)):
            from datetime import datetime, timezone
            try:
                ts = created / 1000 if created > 1e12 else created
                created = datetime.fromtimestamp(ts, tz=timezone.utc).strftime(
                    "%Y-%m-%d %H:%M UTC"
                )
            except Exception as e:
                logger.warning("clerk_timestamp_format_failed", error=str(e))

        entry: dict = {
            "id": user.get("id", ""),
            "name": name,
            "email": primary_email,
            "created_at": created,
        }
        ext_accts = user.get("external_accounts", [])
        if ext_accts and isinstance(ext_accts, list):
            providers = [
                a.get("provider", "") for a in ext_accts
                if isinstance(a, dict)
            ]
            if providers:
                entry["auth"] = ", ".join(p for p in providers if p)
        compact.append(entry)
    return compact

def _build_users_excel(users: list, has_date_filter: bool) -> Path:
    """Create a formatted Excel file with all users and a summary sheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl required for Excel export")

    wb = Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="1E3A5F", end_color="1E3A5F", fill_type="solid",
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    ws_users = wb.active
    ws_users.title = "All Users"
    headers = ["#", "Name", "Email", "Signup Date", "Auth Method"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_users.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    auth_counter: Counter = Counter()
    month_counter: Counter = Counter()

    for row_idx, user in enumerate(users, 2):
        if not isinstance(user, dict):
            continue
        first = (user.get("first_name") or "").strip()
        last = (user.get("last_name") or "").strip()
        name = f"{first} {last}".strip() or "Unknown"
        primary_email = ""
        emails = user.get("email_addresses", [])
        if emails and isinstance(emails, list):
            e = emails[0]
            if isinstance(e, dict):
                primary_email = e.get("email_address", "")

        created_raw = user.get("created_at")
        signup_date = ""
        if isinstance(created_raw, (int, float)):
            ts = created_raw / 1000 if created_raw > 1e12 else created_raw
            try:
                dt = datetime.fromtimestamp(ts, tz=timezone.utc)
                signup_date = dt.strftime("%Y-%m-%d %H:%M")
                month_counter[dt.strftime("%Y-%m")] += 1
            except Exception:
                signup_date = str(created_raw)

        ext_accts = user.get("external_accounts", [])
        auth_method = "email/password"
        if ext_accts and isinstance(ext_accts, list):
            providers = [
                a.get("provider", "") for a in ext_accts if isinstance(a, dict)
            ]
            if providers:
                auth_method = ", ".join(p for p in providers if p) or "email/password"
        auth_counter[auth_method] += 1

        ws_users.cell(row=row_idx, column=1, value=row_idx - 1)
        ws_users.cell(row=row_idx, column=2, value=name)
        ws_users.cell(row=row_idx, column=3, value=primary_email)
        ws_users.cell(row=row_idx, column=4, value=signup_date)
        ws_users.cell(row=row_idx, column=5, value=auth_method)
        for col in range(1, 6):
            ws_users.cell(row=row_idx, column=col).border = thin_border

    ws_users.column_dimensions["A"].width = 6
    ws_users.column_dimensions["B"].width = 25
    ws_users.column_dimensions["C"].width = 35
    ws_users.column_dimensions["D"].width = 18
    ws_users.column_dimensions["E"].width = 20

    ws_summary = wb.create_sheet("Summary", 0)
    summary_headers = ["Metric", "Value"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill

    ws_summary.cell(row=2, column=1, value="Total Users")
    ws_summary.cell(row=2, column=2, value=len(users))
    ws_summary.cell(row=2, column=2).font = Font(bold=True, size=14)

    row = 4
    ws_summary.cell(row=row, column=1, value="Auth Method Breakdown")
    ws_summary.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1
    for method, count in auth_counter.most_common():
        ws_summary.cell(row=row, column=1, value=method)
        ws_summary.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws_summary.cell(row=row, column=1, value="Monthly Signups")
    ws_summary.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1
    for month in sorted(month_counter.keys()):
        ws_summary.cell(row=row, column=1, value=month)
        ws_summary.cell(row=row, column=2, value=month_counter[month])
        row += 1

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    tmp_dir = Path(tempfile.mkdtemp())
    file_path = tmp_dir / "Clerk_Users_Full_Export.xlsx"
    wb.save(str(file_path))
    return file_path


async def clerk_get_user_stats(api_key: str, query_params: Dict = None) -> Dict:
    """Compute aggregate stats across all users without returning full user data.

    Paginates through every user but only tracks counts, avoiding
    the need to send 3000+ user objects back to the model.
    """
    params = dict(query_params or {})
    params.setdefault("limit", 500)
    params["offset"] = 0

    total = 0
    auth_methods: dict[str, int] = {}
    monthly_signups: dict[str, int] = {}

    while True:
        page = await _make_request("GET", "users", api_key, params=params)
        if not isinstance(page, list) or len(page) == 0:
            break
        total += len(page)
        for u in page:
            ext = u.get("external_accounts", [])
            if ext and isinstance(ext, list):
                providers = {
                    a.get("provider", "unknown")
                    for a in ext if isinstance(a, dict)
                }
                for p in providers:
                    auth_methods[p] = auth_methods.get(p, 0) + 1
            else:
                auth_methods["email_password"] = auth_methods.get("email_password", 0) + 1

            created = u.get("created_at")
            if isinstance(created, (int, float)):
                from datetime import datetime, timezone
                try:
                    ts = created / 1000 if created > 1e12 else created
                    month_key = datetime.fromtimestamp(
                        ts, tz=timezone.utc
                    ).strftime("%Y-%m")
                    monthly_signups[month_key] = monthly_signups.get(month_key, 0) + 1
                except Exception as e:
                    logger.warning("clerk_stats_timestamp_parse_failed", error=str(e))

        if len(page) < int(params["limit"]):
            break
        params["offset"] = int(params["offset"]) + len(page)

    return {
        "total_users": total,
        "auth_method_breakdown": dict(sorted(
            auth_methods.items(), key=lambda x: -x[1]
        )),
        "monthly_signups": dict(sorted(monthly_signups.items())),
    }


async def clerk_get_user(api_key: str, user_id: str) -> Dict:
    """
    Retrieves a specific user by their ID.
    """
    return await _make_request("GET", f"users/{user_id}", api_key)

async def clerk_create_user(api_key: str, user_data: Dict) -> Dict:
    """
    Creates a new user in your Clerk instance.
    """
    return await _make_request("POST", "users", api_key, json_data=user_data)

async def clerk_update_user(api_key: str, user_id: str, user_data: Dict) -> Dict:
    """
    Updates an existing user's information.
    """
    return await _make_request("PATCH", f"users/{user_id}", api_key, json_data=user_data)

async def clerk_delete_user(api_key: str, user_id: str) -> Dict:
    """
    Deletes a user from your Clerk instance.
    """
    return await _make_request("DELETE", f"users/{user_id}", api_key)

async def clerk_ban_user(api_key: str, user_id: str) -> Dict:
    """
    Bans a user, preventing them from signing in.
    """
    return await _make_request("POST", f"users/{user_id}/ban", api_key)

async def clerk_unban_user(api_key: str, user_id: str) -> Dict:
    """
    Unbans a user, allowing them to sign in again.
    """
    return await _make_request("POST", f"users/{user_id}/unban", api_key)

# --- Organization Management ---

async def clerk_list_organizations(api_key: str, query_params: Dict = None) -> Dict:
    """
    Lists all organizations in your Clerk instance. Supports filtering and pagination.
    """
    return await _make_request("GET", "organizations", api_key, params=query_params)

async def clerk_get_organization(api_key: str, organization_id: str) -> Dict:
    """
    Retrieves a specific organization by its ID.
    """
    return await _make_request("GET", f"organizations/{organization_id}", api_key)

async def clerk_create_organization(api_key: str, organization_data: Dict) -> Dict:
    """
    Creates a new organization.
    """
    return await _make_request("POST", "organizations", api_key, json_data=organization_data)

async def clerk_update_organization(api_key: str, organization_id: str, organization_data: Dict) -> Dict:
    """
    Updates an existing organization's information.
    """
    return await _make_request("PATCH", f"organizations/{organization_id}", api_key, json_data=organization_data)

async def clerk_delete_organization(api_key: str, organization_id: str) -> Dict:
    """
    Deletes an organization.
    """
    return await _make_request("DELETE", f"organizations/{organization_id}", api_key)

async def clerk_list_organization_memberships(api_key: str, organization_id: str, query_params: Dict = None) -> Dict:
    """
    Lists all memberships for a given organization.
    """
    return await _make_request("GET", f"organizations/{organization_id}/memberships", api_key, params=query_params)

async def clerk_update_organization_membership(api_key: str, organization_id: str, user_id: str, role: str) -> Dict:
    """
    Updates a user's role within an organization.
    """
    return await _make_request("PATCH", f"organizations/{organization_id}/memberships/{user_id}", api_key, json_data={"role": role})

async def clerk_delete_organization_membership(api_key: str, organization_id: str, user_id: str) -> Dict:
    """
    Removes a user from an organization.
    """
    return await _make_request("DELETE", f"organizations/{organization_id}/memberships/{user_id}", api_key)

# --- Webhook Management (Simulated - Clerk's backend API doesn't directly manage webhooks via API) ---
# Note: Clerk's webhooks are typically configured in the dashboard.
# These tools are illustrative for an AI to understand the *concept* of webhooks
# and to potentially interact with a *hypothetical* webhook management API if it existed.
# For a real Clerk integration, webhook events are consumed, not managed via this API.

async def clerk_list_webhooks(api_key: str) -> Dict:
    """
    (Simulated) Lists all configured webhooks. In a real Clerk setup, webhooks are configured
    via the dashboard, and this would represent fetching their configuration.
    """
    return {"error": "Clerk's backend API does not directly manage webhooks via API. Please configure webhooks in the Clerk Dashboard."}

async def clerk_create_webhook(api_key: str, webhook_data: Dict) -> Dict:
    """
    (Simulated) Creates a new webhook endpoint. In a real Clerk setup, webhooks are configured
    via the dashboard. This tool is for conceptual understanding.
    """
    return {"error": "Clerk's backend API does not directly manage webhooks via API. Please configure webhooks in the Clerk Dashboard."}

async def clerk_delete_webhook(api_key: str, webhook_id: str) -> Dict:
    """
    (Simulated) Deletes a webhook endpoint. In a real Clerk setup, webhooks are configured
    via the dashboard. This tool is for conceptual understanding.
    """
    return {"error": "Clerk's backend API does not directly manage webhooks via API. Please configure webhooks in the Clerk Dashboard."}

# --- Session Management ---

async def clerk_list_sessions(api_key: str, query_params: Dict = None) -> Dict:
    """
    Lists all active sessions. Supports filtering by user_id.
    """
    return await _make_request("GET", "sessions", api_key, params=query_params)

async def clerk_get_session(api_key: str, session_id: str) -> Dict:
    """
    Retrieves a specific session by its ID.
    """
    return await _make_request("GET", f"sessions/{session_id}", api_key)

async def clerk_revoke_session(api_key: str, session_id: str) -> Dict:
    """
    Revokes a specific session, logging out the user.
    """
    return await _make_request("POST", f"sessions/{session_id}/revoke", api_key)

# --- Email Address Management ---

async def clerk_list_email_addresses(api_key: str, query_params: Dict = None) -> Dict:
    """
    Lists all email addresses. Supports filtering by user_id.
    """
    return await _make_request("GET", "email_addresses", api_key, params=query_params)

async def clerk_get_email_address(api_key: str, email_address_id: str) -> Dict:
    """
    Retrieves a specific email address by its ID.
    """
    return await _make_request("GET", f"email_addresses/{email_address_id}", api_key)

async def clerk_delete_email_address(api_key: str, email_address_id: str) -> Dict:
    """
    Deletes an email address.
    """
    return await _make_request("DELETE", f"email_addresses/{email_address_id}", api_key)

# --- Phone Number Management ---

async def clerk_list_phone_numbers(api_key: str, query_params: Dict = None) -> Dict:
    """
    Lists all phone numbers. Supports filtering by user_id.
    """
    return await _make_request("GET", "phone_numbers", api_key, params=query_params)

async def clerk_get_phone_number(api_key: str, phone_number_id: str) -> Dict:
    """
    Retrieves a specific phone number by its ID.
    """
    return await _make_request("GET", f"phone_numbers/{phone_number_id}", api_key)

async def clerk_delete_phone_number(api_key: str, phone_number_id: str) -> Dict:
    """
    Deletes a phone number.
    """
    return await _make_request("DELETE", f"phone_numbers/{phone_number_id}", api_key)

# --- Domain Management (for multi-domain applications) ---

async def clerk_list_domains(api_key: str) -> Dict:
    """
    Lists all domains associated with your Clerk instance.
    """
    return await _make_request("GET", "domains", api_key)

async def clerk_get_domain(api_key: str, domain_id: str) -> Dict:
    """
    Retrieves a specific domain by its ID.
    """
    return await _make_request("GET", f"domains/{domain_id}", api_key)

# --- Instance Settings (limited API access) ---

async def clerk_get_instance_settings(api_key: str) -> Dict:
    """
    Retrieves the instance settings for your Clerk application.
    """
    return await _make_request("GET", "instance", api_key)

TOOLS: List[Dict[str, Any]] = [
    {
        "name": "clerk_list_users",
        "description": "Lists users in Clerk. Auto-paginates to get total count. Returns a SAMPLE of 20 users by default. For Excel reports: pass export_excel=true to generate a formatted Excel file with ALL users (2 sheets: raw data + summary with auth breakdown and monthly signups). The file will be created server-side and you just need to upload it. Use created_after_unix_ms to filter by date (Unix ms). Use order_by='-created_at' for newest first.",
        "parameters": {
            "type": "object",
            "properties": {
                "query_params": {
                    "type": "object",
                    "description": "Optional query parameters for filtering and pagination.",
                    "properties": {
                        "email_address": {"type": "string", "description": "Filter by user's email address."},
                        "phone_number": {"type": "string", "description": "Filter by user's phone number."},
                        "external_id": {"type": "string", "description": "Filter by user's external ID."},
                        "username": {"type": "string", "description": "Filter by user's username."},
                        "created_after_unix_ms": {"type": "integer", "description": "Only return users created AFTER this Unix timestamp in MILLISECONDS. Example: 1708992000000 for Feb 27 2024. Use this to count signups in a time period."},
                        "created_before_unix_ms": {"type": "integer", "description": "Only return users created BEFORE this Unix timestamp in MILLISECONDS."},
                        "limit": {"type": "integer", "description": "Number of records to return. Max 500. Use 100 for comprehensive listing.", "default": 100},
                        "offset": {"type": "integer", "description": "Number of records to skip.", "default": 0},
                        "order_by": {"type": "string", "description": "Order results by a specific field. Use '-created_at' for newest first, 'created_at' for oldest first."},
                        "return_all": {"type": "string", "description": "Set to 'true' to return ALL users as JSON. May be large for 1000+ users.", "default": "false"},
                        "export_excel": {"type": "string", "description": "Set to 'true' to generate a formatted Excel file with ALL users. Returns a file path you can upload to Slack. Preferred for reports and exports.", "default": "false"},
                    }
                }
            }
        }
    },
    {
        "name": "clerk_get_user_stats",
        "description": "Get aggregate statistics about all Clerk users. Returns total count, breakdown by authentication method (Google OAuth vs email/password vs others), and monthly signup counts. Use this instead of clerk_list_users when you need aggregate data or breakdowns. Much faster than paginating through all users manually.",
        "parameters": {
            "type": "object",
            "properties": {
                "query_params": {
                    "type": "object",
                    "description": "Optional query parameters for filtering.",
                    "properties": {}
                }
            }
        }
    },
    {
        "name": "clerk_get_user",
        "description": "Retrieves a specific user by their ID.",
        "parameters": {
            "type": "object",
            "properties": {
                "user_id": {"type": "string", "description": "The ID of the user to retrieve.", "pattern": "^user_[a-zA-Z0-9]+$"}
            },
            "required": ["user_id"]
        }
    },
    {
        "name": "clerk_create_user",
        "description": "Creates a new user in your Clerk instance. Requires at least one of `email_address`, `phone_number`, or `username`.",
        "parameters": {
            "type": "object",
            "properties": {
                "user_data": {
                    "type": "object",
                    "description": "Data for the new user.",
                    "properties": {
                        "email_address": {"type": "array", "items": {"type": "string", "format": "email"}, "description": "List of email addresses for the user."},
                        "phone_number": {"type": "array", "items": {"type": "string"}, "description": "List of phone numbers for the user."},
                        "username": {"type": "string", "description": "Username for the user."},
                        "password": {"type": "string", "description": "Password for the user (if not using passwordless)."},
                        "first_name": {"type": "string", "description": "First name of the user."},
                        "last_name": {"type": "string", "description": "Last name of the user."},
                        "external_id": {"type": "string", "description": "An external ID for the user."},
                        "public_metadata": {"type": "object", "description": "Public metadata for the user."},
                        "private_metadata": {"type": "object", "description": "Private metadata for the user."},
                        "unsafe_metadata": {"type": "object", "description": "Unsafe metadata for the user."},
                        "skip_password_checks": {"type": "boolean", "description": "Whether to skip password strength checks.", "default": False},
                        "skip_email_verification": {"type": "boolean", "description": "Whether to skip email verification for new email addresses.", "default": False},
                        "skip_phone_number_verification": {"type": "boolean", "description": "Whether to skip phone number verification for new phone numbers.", "default": False},
                    },
                    "minProperties": 1,
                    "oneOf": [
                        {"required": ["email_address"]},
                        {"required": ["phone_number"]},
                        {"required": ["username"]}
                    ]
                }
            },
            "required": ["user_data"]
        }
    },
    {
        "name": "clerk_update_user",
        "description": "Updates an existing user's information. Only provide fields that need to be updated.",
        "parameters": {
            "type": "object",
            "properties": {
                "user_id": {"type": "string", "description": "The ID of the user to update.", "pattern": "^user_[a-zA-Z0-9]+$"},
                "user_data": {
                    "type": "object",
                    "description": "Data to update for the user.",
                    "properties": {
                        "first_name": {"type": "string", "description": "New first name."},
                        "last_name": {"type": "string", "description": "New last name."},
                        "username": {"type": "string", "description": "New username."},
                        "password": {"type": "string", "description": "New password."},
                        "profile_image_url": {"type": "string", "format": "uri", "description": "New profile image URL."},
                        "public_metadata": {"type": "object", "description": "Public metadata to merge."},
                        "private_metadata": {"type": "object", "description": "Private metadata to merge."},
                        "unsafe_metadata": {"type": "object", "description": "Unsafe metadata to merge."},
                        "external_id": {"type": "string", "description": "New external ID."},
                        "primary_email_address_id": {"type": "string", "description": "ID of the primary email address."},
                        "primary_phone_number_id": {"type": "string", "description": "ID of the primary phone number."},
                        "primary_web3_wallet_id": {"type": "string", "description": "ID of the primary web3 wallet."},
                        "totp_enabled": {"type": "boolean", "description": "Whether TOTP is enabled for the user."},
                        "backup_code_enabled": {"type": "boolean", "description": "Whether backup codes are enabled for the user."},
                    },
                    "minProperties": 1
                }
            },
            "required": ["user_id", "user_data"]
        }
    },
    {
        "name": "clerk_delete_user",
        "description": "Deletes a user from your Clerk instance. This action is irreversible.",
        "parameters": {
            "type": "object",
            "properties": {
                "user_id": {"type": "string", "description": "The ID of the user to delete.", "pattern": "^user_[a-zA-Z0-9]+$"}
            },
            "required": ["user_id"]
        }
    },
    {
        "name": "clerk_ban_user",
        "description": "Bans a user, preventing them from signing in. This does not delete the user.",
        "parameters": {
            "type": "object",
            "properties": {
                "user_id": {"type": "string", "description": "The ID of the user to ban.", "pattern": "^user_[a-zA-Z0-9]+$"}
            },
            "required": ["user_id"]
        }
    },
    {
        "name": "clerk_unban_user",
        "description": "Unbans a user, allowing them to sign in again.",
        "parameters": {
            "type": "object",
            "properties": {
                "user_id": {"type": "string", "description": "The ID of the user to unban.", "pattern": "^user_[a-zA-Z0-9]+$"}
            },
            "required": ["user_id"]
        }
    },
    {
        "name": "clerk_list_organizations",
        "description": "Lists all organizations in your Clerk instance. Supports filtering by `query` (name or slug) and pagination with `limit` and `offset`.",
        "parameters": {
            "type": "object",
            "properties": {
                "query_params": {
                    "type": "object",
                    "description": "Optional query parameters for filtering and pagination.",
                    "properties": {
                        "query": {"type": "string", "description": "Filter organizations by name or slug."},
                        "limit": {"type": "integer", "description": "Number of records to return. Max 500.", "default": 10},
                        "offset": {"type": "integer", "description": "Number of records to skip.", "default": 0},
                        "order_by": {"type": "string", "description": "Order results by a specific field (e.g., 'created_at', '-created_at')."},
                    }
                }
            }
        }
    },
    {
        "name": "clerk_get_organization",
        "description": "Retrieves a specific organization by its ID.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "The ID of the organization to retrieve.", "pattern": "^org_[a-zA-Z0-9]+$"}
            },
            "required": ["organization_id"]
        }
    },
    {
        "name": "clerk_create_organization",
        "description": "Creates a new organization.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_data": {
                    "type": "object",
                    "description": "Data for the new organization.",
                    "properties": {
                        "name": {"type": "string", "description": "The name of the organization.", "minLength": 1},
                        "slug": {"type": "string", "description": "The slug of the organization (unique identifier)."},
                        "created_by": {"type": "string", "description": "The ID of the user creating the organization.", "pattern": "^user_[a-zA-Z0-9]+$"},
                        "public_metadata": {"type": "object", "description": "Public metadata for the organization."},
                        "private_metadata": {"type": "object", "description": "Private metadata for the organization."},
                        "unsafe_metadata": {"type": "object", "description": "Unsafe metadata for the organization."},
                        "image_url": {"type": "string", "format": "uri", "description": "URL of the organization's image."},
                    },
                    "required": ["name", "created_by"]
                }
            },
            "required": ["organization_data"]
        }
    },
    {
        "name": "clerk_update_organization",
        "description": "Updates an existing organization's information. Only provide fields that need to be updated.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "The ID of the organization to update.", "pattern": "^org_[a-zA-Z0-9]+$"},
                "organization_data": {
                    "type": "object",
                    "description": "Data to update for the organization.",
                    "properties": {
                        "name": {"type": "string", "description": "New name of the organization."},
                        "slug": {"type": "string", "description": "New slug of the organization."},
                        "public_metadata": {"type": "object", "description": "Public metadata to merge."},
                        "private_metadata": {"type": "object", "description": "Private metadata to merge."},
                        "unsafe_metadata": {"type": "object", "description": "Unsafe metadata to merge."},
                        "image_url": {"type": "string", "format": "uri", "description": "New URL of the organization's image."},
                    },
                    "minProperties": 1
                }
            },
            "required": ["organization_id", "organization_data"]
        }
    },
    {
        "name": "clerk_delete_organization",
        "description": "Deletes an organization. This action is irreversible.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "The ID of the organization to delete.", "pattern": "^org_[a-zA-Z0-9]+$"}
            },
            "required": ["organization_id"]
        }
    },
    {
        "name": "clerk_list_organization_memberships",
        "description": "Lists all memberships for a given organization. Supports pagination with `limit` and `offset`.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "The ID of the organization.", "pattern": "^org_[a-zA-Z0-9]+$"},
                "query_params": {
                    "type": "object",
                    "description": "Optional query parameters for pagination.",
                    "properties": {
                        "limit": {"type": "integer", "description": "Number of records to return. Max 500.", "default": 10},
                        "offset": {"type": "integer", "description": "Number of records to skip.", "default": 0},
                    }
                }
            },
            "required": ["organization_id"]
        }
    },
    {
        "name": "clerk_update_organization_membership",
        "description": "Updates a user's role within an organization.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "The ID of the organization.", "pattern": "^org_[a-zA-Z0-9]+$"},
                "user_id": {"type": "string", "description": "The ID of the user whose membership to update.", "pattern": "^user_[a-zA-Z0-9]+$"},
                "role": {"type": "string", "description": "The new role for the user in the organization (e.g., 'admin', 'member').", "enum": ["admin", "member"]}
            },
            "required": ["organization_id", "user_id", "role"]
        }
    },
    {
        "name": "clerk_delete_organization_membership",
        "description": "Removes a user from an organization.",
        "parameters": {
            "type": "object",
            "properties": {
                "organization_id": {"type": "string", "description": "The ID of the organization.", "pattern": "^org_[a-zA-Z0-9]+$"},
                "user_id": {"type": "string", "description": "The ID of the user to remove from the organization.", "pattern": "^user_[a-zA-Z0-9]+$"}
            },
            "required": ["organization_id", "user_id"]
        }
    },
    {
        "name": "clerk_list_sessions",
        "description": "Lists all active sessions. Can filter by `user_id`. Supports pagination with `limit` and `offset`.",
        "parameters": {
            "type": "object",
            "properties": {
                "query_params": {
                    "type": "object",
                    "description": "Optional query parameters for filtering and pagination.",
                    "properties": {
                        "user_id": {"type": "string", "description": "Filter sessions by user ID.", "pattern": "^user_[a-zA-Z0-9]+$"},
                        "client_id": {"type": "string", "description": "Filter sessions by client ID.", "pattern": "^client_[a-zA-Z0-9]+$"},
                        "status": {"type": "string", "description": "Filter sessions by status (e.g., 'active', 'revoked')."},
                        "limit": {"type": "integer", "description": "Number of records to return. Max 500.", "default": 10},
                        "offset": {"type": "integer", "description": "Number of records to skip.", "default": 0},
                    }
                }
            }
        }
    },
    {
        "name": "clerk_get_session",
        "description": "Retrieves a specific session by its ID.",
        "parameters": {
            "type": "object",
            "properties": {
                "session_id": {"type": "string", "description": "The ID of the session to retrieve.", "pattern": "^sess_[a-zA-Z0-9]+$"}
            },
            "required": ["session_id"]
        }
    },
    {
        "name": "clerk_revoke_session",
        "description": "Revokes a specific session, effectively logging out the user associated with it.",
        "parameters": {
            "type": "object",
            "properties": {
                "session_id": {"type": "string", "description": "The ID of the session to revoke.", "pattern": "^sess_[a-zA-Z0-9]+$"}
            },
            "required": ["session_id"]
        }
    },
    {
        "name": "clerk_list_email_addresses",
        "description": "Lists all email addresses. Can filter by `user_id`. Supports pagination with `limit` and `offset`.",
        "parameters": {
            "type": "object",
            "properties": {
                "query_params": {
                    "type": "object",
                    "description": "Optional query parameters for filtering and pagination.",
                    "properties": {
                        "user_id": {"type": "string", "description": "Filter email addresses by user ID.", "pattern": "^user_[a-zA-Z0-9]+$"},
                        "limit": {"type": "integer", "description": "Number of records to return. Max 500.", "default": 10},
                        "offset": {"type": "integer", "description": "Number of records to skip.", "default": 0},
                    }
                }
            }
        }
    },
    {
        "name": "clerk_get_email_address",
        "description": "Retrieves a specific email address by its ID.",
        "parameters": {
            "type": "object",
            "properties": {
                "email_address_id": {"type": "string", "description": "The ID of the email address to retrieve.", "pattern": "^ema_[a-zA-Z0-9]+$"}
            },
            "required": ["email_address_id"]
        }
    },
    {
        "name": "clerk_delete_email_address",
        "description": "Deletes an email address. This action is irreversible.",
        "parameters": {
            "type": "object",
            "properties": {
                "email_address_id": {"type": "string", "description": "The ID of the email address to delete.", "pattern": "^ema_[a-zA-Z0-9]+$"}
            },
            "required": ["email_address_id"]
        }
    },
    {
        "name": "clerk_list_phone_numbers",
        "description": "Lists all phone numbers. Can filter by `user_id`. Supports pagination with `limit` and `offset`.",
        "parameters": {
            "type": "object",
            "properties": {
                "query_params": {
                    "type": "object",
                    "description": "Optional query parameters for filtering and pagination.",
                    "properties": {
                        "user_id": {"type": "string", "description": "Filter phone numbers by user ID.", "pattern": "^user_[a-zA-Z0-9]+$"},
                        "limit": {"type": "integer", "description": "Number of records to return. Max 500.", "default": 10},
                        "offset": {"type": "integer", "description": "Number of records to skip.", "default": 0},
                    }
                }
            }
        }
    },
    {
        "name": "clerk_get_phone_number",
        "description": "Retrieves a specific phone number by its ID.",
        "parameters": {
            "type": "object",
            "properties": {
                "phone_number_id": {"type": "string", "description": "The ID of the phone number to retrieve.", "pattern": "^phn_[a-zA-Z0-9]+$"}
            },
            "required": ["phone_number_id"]
        }
    },
    {
        "name": "clerk_delete_phone_number",
        "description": "Deletes a phone number. This action is irreversible.",
        "parameters": {
            "type": "object",
            "properties": {
                "phone_number_id": {"type": "string", "description": "The ID of the phone number to delete.", "pattern": "^phn_[a-zA-Z0-9]+$"}
            },
            "required": ["phone_number_id"]
        }
    },
    {
        "name": "clerk_list_domains",
        "description": "Lists all domains associated with your Clerk instance.",
        "parameters": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "clerk_get_domain",
        "description": "Retrieves a specific domain by its ID.",
        "parameters": {
            "type": "object",
            "properties": {
                "domain_id": {"type": "string", "description": "The ID of the domain to retrieve.", "pattern": "^dom_[a-zA-Z0-9]+$"}
            },
            "required": ["domain_id"]
        }
    },
    {
        "name": "clerk_get_instance_settings",
        "description": "Retrieves the instance settings for your Clerk application.",
        "parameters": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "clerk_list_webhooks",
        "description": "(Simulated) Lists all configured webhooks. In a real Clerk setup, webhooks are configured via the dashboard, and this tool is for conceptual understanding.",
        "parameters": {
            "type": "object",
            "properties": {}
        }
    },
    {
        "name": "clerk_create_webhook",
        "description": "(Simulated) Creates a new webhook endpoint. In a real Clerk setup, webhooks are configured via the dashboard. This tool is for conceptual understanding.",
        "parameters": {
            "type": "object",
            "properties": {
                "webhook_data": {
                    "type": "object",
                    "description": "Data for the new webhook.",
                    "properties": {
                        "url": {"type": "string", "format": "uri", "description": "The URL where webhook events will be sent."},
                        "events": {"type": "array", "items": {"type": "string"}, "description": "List of event types to subscribe to (e.g., 'user.created', 'organization.created')."},
                        "secret": {"type": "string", "description": "Optional secret for signing webhook payloads."},
                    },
                    "required": ["url", "events"]
                }
            },
            "required": ["webhook_data"]
        }
    },
    {
        "name": "clerk_delete_webhook",
        "description": "(Simulated) Deletes a webhook endpoint. In a real Clerk setup, webhooks are configured via the dashboard. This tool is for conceptual understanding.",
        "parameters": {
            "type": "object",
            "properties": {
                "webhook_id": {"type": "string", "description": "The ID of the webhook to delete."}
            },
            "required": ["webhook_id"]
        }
    }
]

async def execute(tool_name: str, args: Dict, api_key: str) -> Dict:
    """
    Dispatches to the appropriate Clerk API call based on the tool_name.
    """
    if tool_name == "clerk_list_users":
        return await clerk_list_users(api_key, **args)
    elif tool_name == "clerk_get_user_stats":
        return await clerk_get_user_stats(api_key, **args)
    elif tool_name == "clerk_get_user":
        return await clerk_get_user(api_key, **args)
    elif tool_name == "clerk_create_user":
        return await clerk_create_user(api_key, **args)
    elif tool_name == "clerk_update_user":
        return await clerk_update_user(api_key, **args)
    elif tool_name == "clerk_delete_user":
        return await clerk_delete_user(api_key, **args)
    elif tool_name == "clerk_ban_user":
        return await clerk_ban_user(api_key, **args)
    elif tool_name == "clerk_unban_user":
        return await clerk_unban_user(api_key, **args)
    elif tool_name == "clerk_list_organizations":
        return await clerk_list_organizations(api_key, **args)
    elif tool_name == "clerk_get_organization":
        return await clerk_get_organization(api_key, **args)
    elif tool_name == "clerk_create_organization":
        return await clerk_create_organization(api_key, **args)
    elif tool_name == "clerk_update_organization":
        return await clerk_update_organization(api_key, **args)
    elif tool_name == "clerk_delete_organization":
        return await clerk_delete_organization(api_key, **args)
    elif tool_name == "clerk_list_organization_memberships":
        return await clerk_list_organization_memberships(api_key, **args)
    elif tool_name == "clerk_update_organization_membership":
        return await clerk_update_organization_membership(api_key, **args)
    elif tool_name == "clerk_delete_organization_membership":
        return await clerk_delete_organization_membership(api_key, **args)
    elif tool_name == "clerk_list_sessions":
        return await clerk_list_sessions(api_key, **args)
    elif tool_name == "clerk_get_session":
        return await clerk_get_session(api_key, **args)
    elif tool_name == "clerk_revoke_session":
        return await clerk_revoke_session(api_key, **args)
    elif tool_name == "clerk_list_email_addresses":
        return await clerk_list_email_addresses(api_key, **args)
    elif tool_name == "clerk_get_email_address":
        return await clerk_get_email_address(api_key, **args)
    elif tool_name == "clerk_delete_email_address":
        return await clerk_delete_email_address(api_key, **args)
    elif tool_name == "clerk_list_phone_numbers":
        return await clerk_list_phone_numbers(api_key, **args)
    elif tool_name == "clerk_get_phone_number":
        return await clerk_get_phone_number(api_key, **args)
    elif tool_name == "clerk_delete_phone_number":
        return await clerk_delete_phone_number(api_key, **args)
    elif tool_name == "clerk_list_domains":
        return await clerk_list_domains(api_key, **args)
    elif tool_name == "clerk_get_domain":
        return await clerk_get_domain(api_key, **args)
    elif tool_name == "clerk_get_instance_settings":
        return await clerk_get_instance_settings(api_key, **args)
    elif tool_name == "clerk_list_webhooks":
        return await clerk_list_webhooks(api_key, **args)
    elif tool_name == "clerk_create_webhook":
        return await clerk_create_webhook(api_key, **args)
    elif tool_name == "clerk_delete_webhook":
        return await clerk_delete_webhook(api_key, **args)
    else:
        return {"error": f"Tool '{tool_name}' not found."}